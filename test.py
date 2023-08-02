import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

def save_employe():
    first_name = entry_first_name.get()
    last_name = entry_last_name.get()
    sexe = combo_sexe.get()
    age = int(spin_age.get())
    departement = entry_departement.get()
    adresse = entry_adresse.get()
    tel = entry_tel.get()
    email = entry_email.get()

    if first_name and last_name and sexe and age and departement and adresse and tel and email:
        if email_exists(email):
            messagebox.showerror("Erreur", "L'email existe déjà.")
            return

        wb = load_workbook("employes.xlsx") if "employes.xlsx" in os.listdir() else Workbook()
        sheet = wb.active if "Sheet" in wb.sheetnames else wb.create_sheet()

        if sheet.max_row == 1:
            headers = ["First Name", "Last Name", "Sexe", "Age", "Departement", "Adresse", "Telephone", "Email"]
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                sheet[f"{col_letter}1"] = header

        next_row = sheet.max_row + 1
        employe_data = [first_name, last_name, sexe, age, departement, adresse, tel, email]
        for col_num, value in enumerate(employe_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{next_row}"] = value

        wb.save("employes.xlsx")
        messagebox.showinfo("Succès", "L'employé a été enregistré avec succès.")
        clear_entries()  
    else:
        messagebox.showerror("Erreur", "Veuillez saisir toutes les informations de l'employé.")

def email_exists(email):
    wb = load_workbook("employes.xlsx") if "employes.xlsx" in os.listdir() else None
    if wb:
        sheet = wb.active if "Sheet" in wb.sheetnames else None
        if sheet:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[7] == email:
                    return True
    return False

def clear_entries():
    entry_first_name.delete(0, tk.END)
    entry_last_name.delete(0, tk.END)
    combo_sexe.set("Homme")
    spin_age.delete(0, tk.END)
    entry_departement.delete(0, tk.END)
    entry_adresse.delete(0, tk.END)
    entry_tel.delete(0, tk.END)
    entry_email.delete(0, tk.END)

window = tk.Tk()
window.title("Saisie des informations de l'employé")
window.geometry("800x300")

user_frame = tk.LabelFrame(window, text="USER", padx=10, pady=10, labelanchor="n", font=("Helvetica", 12, "bold"))
user_frame.pack(padx=10, pady=10, anchor="center")

entry_style = {"width": 30, "font": ("Helvetica", 12)}

label_first_name = tk.Label(user_frame, text="First Name:", font=("Helvetica", 12))
label_first_name.grid(row=0, column=0, padx=5, pady=5)
entry_first_name = tk.Entry(user_frame, **entry_style)
entry_first_name.grid(row=0, column=1, padx=5, pady=5)

label_last_name = tk.Label(user_frame, text="Last Name:", font=("Helvetica", 12))
label_last_name.grid(row=0, column=2, padx=5, pady=5)
entry_last_name = tk.Entry(user_frame, **entry_style)
entry_last_name.grid(row=0, column=3, padx=5, pady=5)

label_sexe = tk.Label(user_frame, text="Sexe:", font=("Helvetica", 12))
label_sexe.grid(row=1, column=0, padx=5, pady=5)
combo_sexe = tk.StringVar()
combo_sexe.set("Homme")
entry_sexe = tk.OptionMenu(user_frame, combo_sexe, "Homme", "Femme")
entry_sexe.grid(row=1, column=1, padx=5, pady=5)

label_age = tk.Label(user_frame, text="Age:", font=("Helvetica", 12))
label_age.grid(row=1, column=2, padx=5, pady=5)
spin_age = tk.Spinbox(user_frame, from_=18, to=110, **entry_style)
spin_age.grid(row=1, column=3, padx=5, pady=5)

label_departement = tk.Label(user_frame, text="Département:", font=("Helvetica", 12))
label_departement.grid(row=2, column=0, padx=5, pady=5)
entry_departement = tk.Entry(user_frame, **entry_style)
entry_departement.grid(row=2, column=1, padx=5, pady=5, columnspan=3)

label_adresse = tk.Label(user_frame, text="Adresse:", font=("Helvetica", 12))
label_adresse.grid(row=3, column=0, padx=5, pady=5)
entry_adresse = tk.Entry(user_frame, **entry_style)
entry_adresse.grid(row=3, column=1, padx=5, pady=5, columnspan=3)

label_tel = tk.Label(user_frame, text="Téléphone:", font=("Helvetica", 12))
label_tel.grid(row=4, column=0, padx=5, pady=5)
entry_tel = tk.Entry(user_frame, **entry_style)
entry_tel.grid(row=4, column=1, padx=5, pady=5)

label_email = tk.Label(user_frame, text="Email:", font=("Helvetica", 12))
label_email.grid(row=4, column=2, padx=5, pady=5)
entry_email = tk.Entry(user_frame, **entry_style)
entry_email.grid(row=4, column=3, padx=5, pady=5)

button_save = tk.Button(window, text="Enregistrer", command=save_employe, bg="#2e91a5", fg="black", font=("Helvetica", 12))
button_save.pack(pady=10)

window.mainloop()
